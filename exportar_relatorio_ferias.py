import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO
from datetime import datetime
from database.models import Ferias, Funcionario

def exportar_relatorio_ferias(session, empresa_nome):
    """
    Exporta relatório de férias para Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Férias"

    # Configurar tamanho da página para A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Relatório de Férias - {empresa_nome}"
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Cabeçalhos
    headers = ['Nome', 'Função', 'Horário', 'Início das Férias', 'Fim das Férias', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Configurar largura das colunas
    ws.column_dimensions['A'].width = 30  # Nome
    ws.column_dimensions['B'].width = 20  # Função
    ws.column_dimensions['C'].width = 15  # Horário
    ws.column_dimensions['D'].width = 15  # Início
    ws.column_dimensions['E'].width = 15  # Fim
    ws.column_dimensions['F'].width = 15  # Status

    # Query para buscar todos os registros de férias com informações do funcionário
    ferias_records = (
        session.query(
            Ferias,
            Funcionario.nome,
            Funcionario.funcao,
            Funcionario.horario_turno
        )
        .join(Funcionario, Ferias.funcionario_id == Funcionario.id)
        .order_by(Ferias.data_inicio.desc())  # Ordenar por data de início, mais recente primeiro
        .all()
    )

    # Adicionar dados ao Excel
    row = 4
    for ferias, nome, funcao, horario in ferias_records:
        ws.cell(row=row, column=1, value=nome)
        ws.cell(row=row, column=2, value=funcao)
        ws.cell(row=row, column=3, value=horario)
        ws.cell(row=row, column=4, value=ferias.data_inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=ferias.data_fim.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=6, value='Em andamento' if ferias.ativa else 'Concluído')
        
        # Aplicar estilos
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = cell_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colorir a célula de status
            if col == 6:  # Coluna de status
                if ferias.ativa:
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Amarelo claro
                else:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Verde claro
        
        row += 1

    # Salvar o arquivo Excel
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def adicionar_botao_exportacao_ferias(session, empresa_nome):
    """
    Adiciona botão para exportar relatório de férias
    """
    if st.button("Exportar Relatório de Férias"):
        try:
            excel_file = exportar_relatorio_ferias(session, empresa_nome)
            
            st.download_button(
                label="Baixar Relatório de Férias",
                data=excel_file,
                file_name=f"relatorio_ferias_{empresa_nome}_{datetime.now().strftime('%B_%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Erro ao gerar relatório: {str(e)}")
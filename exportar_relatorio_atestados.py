import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from database.models import Atestado, Funcionario

def exportar_relatorio_atestados(session, empresa_nome):
    """
    Exporta relatório de atestados para Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Atestados"

    # Configurar tamanho da página para A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Relatório de Atestados - {empresa_nome}"
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Cabeçalhos
    headers = ['Nome', 'Função', 'Data Início', 'Data Fim', 'Dias', 'Motivo', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Configurar largura das colunas
    ws.column_dimensions['A'].width = 30  # Nome
    ws.column_dimensions['B'].width = 20  # Função
    ws.column_dimensions['C'].width = 15  # Data Início
    ws.column_dimensions['D'].width = 15  # Data Fim
    ws.column_dimensions['E'].width = 10  # Dias
    ws.column_dimensions['F'].width = 40  # Motivo
    ws.column_dimensions['G'].width = 15  # Status

    # Query para buscar todos os registros de atestados com informações do funcionário
    atestados_records = (
        session.query(
            Atestado,
            Funcionario.nome,
            Funcionario.funcao
        )
        .join(Funcionario, Atestado.funcionario_id == Funcionario.id)
        .order_by(Atestado.data_inicio.desc())  # Ordenar por data de início, mais recente primeiro
        .all()
    )

    # Adicionar dados ao Excel
    row = 4
    for atestado, nome, funcao in atestados_records:
        ws.cell(row=row, column=1, value=nome)
        ws.cell(row=row, column=2, value=funcao)
        ws.cell(row=row, column=3, value=atestado.data_inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=4, value=atestado.data_fim.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=atestado.dias)
        ws.cell(row=row, column=6, value=atestado.motivo)
        ws.cell(row=row, column=7, value='Em andamento' if atestado.ativo else 'Concluído')
        
        # Aplicar estilos
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = cell_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colorir a célula de status
            if col == 7:  # Coluna de status
                if atestado.ativo:
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Amarelo claro
                else:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Verde claro
            
            # Alinhar motivo à esquerda
            if col == 6:  # Coluna do motivo
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        row += 1

    # Salvar o arquivo Excel
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
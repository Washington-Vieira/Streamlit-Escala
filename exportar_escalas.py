import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

def exportar_escalas_para_excel(df_escala_final, df_folguistas, empresa_nome, mes_ano):
    wb = Workbook()
    ws = wb.active
    ws.title = "Escala de Trabalho"

    # Configurar tamanho da página para A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Título
    ws.merge_cells('A1:AF1')
    ws['A1'] = f"Escala de Trabalho - {empresa_nome} - {mes_ano}"
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Mesclar células da linha 2
    ws.merge_cells('A2:AF2')
    
    # Ajustar o alinhamento da célula mesclada (opcional)
    merged_cell = ws['A2']
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Função para configurar dimensões e estilos das células
    def configurar_celulas(worksheet, df, linha_inicio=4, linha_inicio_folguistas=None, 
                           linha_titulo_folguistas=None, linha_cabecalho_folguistas=None):
        # Configurar largura das colunas
        for col in worksheet.columns:
            col_letter = get_column_letter(col[0].column)
            if col[0].column == 1:  # Coluna dos nomes
                worksheet.column_dimensions[col_letter].width = 20
            else:  # Colunas dos dias
                worksheet.column_dimensions[col_letter].width = 6

        # Configurar altura das linhas
        for row in range(linha_inicio, worksheet.max_row + 1):
            if linha_inicio_folguistas is None or linha_titulo_folguistas is None or linha_cabecalho_folguistas is None or \
               row not in [linha_inicio_folguistas, linha_titulo_folguistas, linha_cabecalho_folguistas]:
                worksheet.row_dimensions[row].height = 50.25

        # Aplicar estilos
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if cell.row == 3 or cell.row == linha_cabecalho_folguistas:
                    cell.font = Font(size=10, bold=True)
                else:
                    cell.font = cell_font

        # Ajustar o tamanho da fonte da linha 3 para 10
        for col in worksheet.iter_cols(min_row=3, max_row=3):
            for cell in col:
                cell.font = Font(size=10, bold=True)

        # Aplicar formatação condicional
        verde = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        laranja = PatternFill(start_color='FF6400', end_color='FF6400', fill_type='solid')
        azul = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')

        # Aplicar regra para "folga"
        worksheet.conditional_formatting.add(f'B4:{get_column_letter(worksheet.max_column)}{worksheet.max_row}',
                                             CellIsRule(operator='equal', formula=['"folga"'], fill=verde))

        # Aplicar regra para "folga (domingo)"
        worksheet.conditional_formatting.add(f'B4:{get_column_letter(worksheet.max_column)}{worksheet.max_row}',
                                             CellIsRule(operator='equal', formula=['"folga (domingo)"'], fill=laranja))

        # Identificar e pintar os sábados de azul
        for col in range(2, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=3, column=col).value
            if cell_value and isinstance(cell_value, str):
                if 'sáb' in cell_value.lower():
                    for row in range(4, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = azul

    # Remover bordas internas da linha 2
    no_border = Border(left=Side(style=None), 
                       right=Side(style=None), 
                       top=Side(style=None), 
                       bottom=Side(style=None))

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        if col == 1:
            # Manter borda esquerda para a primeira coluna
            cell.border = Border(left=Side(style='thin'), right=Side(style=None), 
                                 top=Side(style=None), bottom=Side(style=None))
        elif col == ws.max_column:
            # Manter borda direita para a última coluna
            cell.border = Border(left=Side(style=None), right=Side(style='thin'), 
                                 top=Side(style=None), bottom=Side(style=None))
        else:
            # Remover todas as bordas para as colunas do meio
            cell.border = no_border

    # Ajustar o tamanho da fonte da linha 3 para 10 e manter em negrito
    for col in ws.iter_cols(min_row=3, max_row=3):
        for cell in col:
            cell.font = Font(size=10, bold=True)

    # Escala de trabalho
    if not df_escala_final.empty:
        # Cabeçalho da escala de trabalho
        for col, header in enumerate(df_escala_final.columns, start=1):
            ws.cell(row=3, column=col, value=header)

        # Dados da escala de trabalho
        for row, data in enumerate(df_escala_final.itertuples(index=False), start=4):
            for col, value in enumerate(data, start=1):
                ws.cell(row=row, column=col, value=value)

        configurar_celulas(ws, df_escala_final)
    else:
        ws.cell(row=3, column=1, value="Nenhuma escala de trabalho disponível")

    # Adicionar escala de folguistas na mesma página
    ultima_linha_escala = ws.max_row
    linha_inicio_folguistas = ultima_linha_escala + 1

    # Primeira linha abaixo da escala principal (mesclada e com altura de 15)
    ws.merge_cells(f'A{linha_inicio_folguistas}:AF{linha_inicio_folguistas}')
    ws.row_dimensions[linha_inicio_folguistas].height = 15

    # Segunda linha com o título da escala de folguistas (mesclada e com altura de 32)
    linha_titulo_folguistas = linha_inicio_folguistas + 1
    ws.merge_cells(f'A{linha_titulo_folguistas}:AF{linha_titulo_folguistas}')
    ws[f'A{linha_titulo_folguistas}'] = "Escala de Folguistas"
    ws[f'A{linha_titulo_folguistas}'].font = titulo_font
    ws[f'A{linha_titulo_folguistas}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[linha_titulo_folguistas].height = 32

    if not df_folguistas.empty:
        # Cabeçalho da escala de folguistas
        linha_cabecalho_folguistas = linha_titulo_folguistas + 1
        
        # Adicionar cabeçalhos
        for col, header in enumerate(df_folguistas.columns, start=1):
            cell = ws.cell(row=linha_cabecalho_folguistas, column=col, value=header)
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Definir altura da linha de cabeçalho dos folguistas para 15
        ws.row_dimensions[linha_cabecalho_folguistas].height = 15

        # Dados da escala de folguistas
        for row_idx, row in df_folguistas.iterrows():
            excel_row = linha_cabecalho_folguistas + 1 + row_idx
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=excel_row, column=col_idx + 1)
                cell.value = str(value) if pd.notna(value) else ''
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Aplicar formatação condicional para folgas
                if isinstance(value, str) and 'folga' in value.lower():
                    if 'domingo' in value.lower():
                        cell.fill = PatternFill(start_color='FF6400', end_color='FF6400', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')

        configurar_celulas(ws, df_folguistas, linha_inicio=linha_cabecalho_folguistas, 
                           linha_inicio_folguistas=linha_inicio_folguistas, 
                           linha_titulo_folguistas=linha_titulo_folguistas,
                           linha_cabecalho_folguistas=linha_cabecalho_folguistas)
    else:
        ws.cell(row=linha_titulo_folguistas + 1, column=1, value="Nenhuma escala de folguistas disponível")

    # Salvar o arquivo Excel
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

# Função para adicionar o botão de exportação nas páginas
def adicionar_botao_exportacao(df_escala_final, df_folguistas, empresa_nome):
    if st.button('Exportar Escalas para Excel'):
        # Obter o mês e ano atual se não houver coluna 'Data'
        mes_ano = datetime.now().strftime('%B %Y')
        
        # Se houver uma coluna 'Data', use-a para obter o mês e ano
        if 'Data' in df_escala_final.columns and not df_escala_final['Data'].empty:
            mes_ano = pd.to_datetime(df_escala_final['Data'].iloc[0]).strftime('%B %Y')
        
        excel_file = exportar_escalas_para_excel(df_escala_final, df_folguistas, empresa_nome, mes_ano)
        
        st.download_button(
            label="Baixar Escalas em Excel",
            data=excel_file,
            file_name=f"escalas_{empresa_nome}_{mes_ano}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
